# -*- coding: utf-8 -*-
"""
@File    : excel_handler.py
@Time    : 2026/7/29 16:34
@Author  : Your Name
@Version : 1.0
@Desc    : 
"""
from openpyxl import load_workbook
from common.log_util import log
from typing import List, Dict

class ExcelHandler:
    @staticmethod
    def read_case(file_path: str, sheet_name: str) -> List[Dict]:
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                raise Exception(f"工作表 {sheet_name} 不存在")
            ws = wb[sheet_name]
            headers = [cell.value.strip() if cell.value else "" for cell in ws[1]]
            case_list = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for idx, val in enumerate(row):
                    k = headers[idx]
                    if val is None or str(val).strip() == "":
                        row_dict[k] = None
                    else:
                        row_dict[k] = str(val).strip()
                case_list.append(row_dict)
            wb.close()
            log.info(f"读取Excel用例成功，共{len(case_list)}条")
            return case_list
        except Exception as e:
            log.error(f"Excel读取失败：{e}")
            raise e

    @staticmethod
    def filter_by_tag(case_list: List[Dict], tag: str):
        return [case for case in case_list if case.get("tag") == tag]

    @staticmethod
    def get_param_pair(case_list: List[Dict]):
        return case_list

excel_handler = ExcelHandler()